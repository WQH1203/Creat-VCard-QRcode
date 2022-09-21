from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

# 创建excel
# wb = workbook

# 读取excel
wb = load_workbook('../dados.xlsx')

# 读取工作表
# st = wb.active
st = wb['Folha2']

# 修改单元格的值
# st['B7'].value = 'G'
# print(st['B7'].value)

# 增加工作表
# wb.create_sheet('Folha3')

# 列出工作表
# print(wb.sheetnames)

# 修改工资表名称
# st.title = 'People'

# 横排增加
# st.append([1, 2, 3, 4])
# st.append([5, 6, 7, 8])

# 读取范围资料
# for row in range(1, 3):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         st[char + str(row)].value = char + str(row)

# 合并单元格
# st.merge_cells('D1:E1')

# 拆分单元格
# st.unmerge_cells('D1:E1')

# 插入横/竖排
# st.insert_rows(1)
# st.insert_cols(1)

# 删除横/竖排
# aux_01 = 3
# st.delete_rows(idx=aux_01)
# st.delete_cols(2)

# 移动单元格
# st.move_range('A1:D2', rows=4, cols=4)

# 获取工作表最大行数和最大列数
# print(st.max_row)
# print(st.max_column)

# 保存excel
wb.save('dados.xlsx')

# import tkinter as tk
# from tkinter import filedialog
# from PIL import Image
#
#
# def open_img():
#     try:
#         global img
#         filepath = filedialog.askopenfilename()  # 打开文件，返回该文件的完整路径
#         filename.set(filepath)
#         img = Image.open(filename.get())
#     except Exception as e:
#         print("您没有选择任何文件", e)
#
#
# def save_png():
#     try:
#         filetypes = [("PNG", "*.png")]
#         # 返回一个 pathname 文件路径字符串，如果取消或者关闭则返回空字符，返回文件如何操作是后续代码的事情，
#         # 该函数知识返回选择文件的文件名字，不具备保存文件的能力
#         filenewpath = filedialog.asksaveasfilename(title='保存文件', filetypes=filetypes, defaultextension='.png')
#         path_var.set(filenewpath)
#         # 保存文件
#         img.save(str(path_var.get()))
#     except Exception as e:
#         print(e)
#
#
# window = tk.Tk()
# window.geometry('400x200+300+300')
#
#
# filename = tk.StringVar()
# path_var = tk.StringVar()
#
# # 定义读取文件的组件
# entry = tk.Entry(window, textvariable=filename)
# entry.grid(row=1, column=0, padx=5, pady=5)
# tk.Button(window, text='选择文件', command=open_img).grid(row=1, column=1, padx=5, pady=5)
#
# # 定义保存文件的组件
# entry1 = tk.Entry(window, textvariable=path_var)
# entry1.grid(row=2, column=0, padx=5, pady=5)
# tk.Button(window, text='保存文件', command=save_png).grid(row=2, column=1, padx=5, pady=5)
# window.mainloop()
