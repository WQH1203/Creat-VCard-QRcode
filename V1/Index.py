from tkinter import *
import tkinter.messagebox
from tkinter import filedialog
import os
import qrcode
from PIL import Image, ImageTk
from openpyxl import workbook, load_workbook
import mysql.connector as myconect

mysql_connection = myconect.connect(host="localhost", user="Antonio", password="20031203zhijie", database="qr_code_db")
cursor = mysql_connection.cursor()

contador = 0
v_excel = os.path.exists('dados.xlsx')

if v_excel:
    wb = load_workbook('dados.xlsx')
    st = wb['People']
else:
    wb = workbook
    st = wb.active
    st.title = 'People'
    title = ['Id', 'Nome']
    st.append(title)


def clear():
    lbl_Id_V['text'] = ''
    txt_Name.delete("1.0", "end")


def messagebox(title_aux, mensagem):
    tkinter.messagebox.showwarning(title_aux, mensagem)


def get_name():
    name = txt_Name.get("0.0", "end").replace("\n", "").replace("\r", "")
    return name


def procurador():
    global contador
    for row in range(2, st.max_row + 1):
        aux01 = st['A' + str(row)].value
        if str(aux01) == str(numEty.get()):
            contador = row
            return row
        else:
            if aux01 is None and (row - 1) != st.max_row:
                if str((row - 1)) == str(numEty.get()):
                    return 0
            else:
                continue


def procurar():
    clear()
    global photo
    if numEty.get().isdigit() and st.max_row > int(numEty.get()) > 0:
        aux_01: int = procurador()
        if aux_01 != 0:
            lbl_Id_V['text'] = st['A' + str(aux_01)].value
            txt_Name.insert(INSERT, st['B' + str(aux_01)].value)
            ver_img = os.path.exists('./IMG/' + numEty.get() + ".png")
            if ver_img:
                img = Image.open('./IMG/%s.png' % numEty.get())
                photo = ImageTk.PhotoImage(img)
                lbl_Fm_Img['image'] = photo
        else:
            messagebox('Error', 'Esse Id não tem imformações')
    else:
        messagebox('Error', 'Id invalido')


def salvar():
    global contador
    if get_name() == '' or get_name() == str(st['B' + str(contador)].value):
        messagebox('Error', 'Indica o nome')
    else:
        st['B' + str(contador)].value = get_name()
        wb.save('dados.xlsx')


def create():
    aux_num01 = 0
    if get_name() == '':
        messagebox('Error', 'Indica o nome')
    else:
        for i in range(2, st.max_row):
            aux_num02 = st['A' + str(i)].value
            if aux_num02 is None:
                st['A' + str(i)].value = i - 1
                st['B' + str(i)].value = get_name()
                aux_num01 = 1
                break
        if aux_num01 == 0:
            st['A' + str(st.max_row + 1)].value = st.max_row
            st['B' + str(st.max_row)].value = get_name()
        wb.save('dados.xlsx')


def delete_data():
    if numEty.get().isdigit() and int(numEty.get()) <= st.max_row:
        aux_01 = procurador()
        if aux_01 != 0:
            st.delete_rows(idx=aux_01)
            if aux_01 != st.max_row:
                st.insert_rows(idx=aux_01)
            messagebox('Delete', "Imformação eleminado")
            wb.save('dados.xlsx')
        else:
            st.delete_rows(idx=st.max_row)
            messagebox('Delete', "Imformação não existe")


vcard = '''BEGIN:VCARD
VERSION:4.0
NOTE:ID:%s
FN:%s
PHOTO;VALUE=URL:%s
END:VCARD'''


def qrc():
    global vcard
    qr = qrcode.QRCode(
        version=5,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=5,
        border=1,
    )
    if numEty.get() <= str(10):
        vcard = vcard % (numEty.get(), get_name(), "https://pigsrockfestival.com/antonio/" + str(numEty.get()) + ".png")
    else:
        vcard = vcard % (numEty.get(), get_name(), None)

    qr.add_data(vcard)
    img = qr.make_image()
    img.save("QR_Image/" + numEty.get() + '.png')

    img_02 = Image.open('QR_Image/%s.png' % numEty.get())
    photo_02 = ImageTk.PhotoImage(img_02)
    lbl_img['image'] = photo_02


def create_qrc():
    if numEty.get() == '':
        messagebox('Erro', "Indica o Id que é para fazer QRcode")
    else:
        if os.path.exists('QR_Image'):
            qrc()
        else:
            os.makedirs("QR_Image")
            qrc()


def create_img():
    if numEty.get() == '':
        messagebox("erro", "Indica o Id")
    else:
        global photo_03, photo_04, img_04
        ver_img = os.path.exists('./IMG/' + numEty.get() + ".png")
        if ver_img:
            img_03 = Image.open('./IMG/%s.png' % numEty.get())
            photo_03 = ImageTk.PhotoImage(img_03)
            lbl_Fm_Img['image'] = photo_03
        else:
            filepath = filedialog.askopenfilename()
            if filepath == '':
                messagebox("erro", "Não seleccionou nenhum documento")
            else:
                img_04 = Image.open(filepath)
                photo_04 = ImageTk.PhotoImage(img_04)
                lbl_Fm_Img['image'] = photo_04
                img_04.save(os.path.join("./IMG/", numEty.get() + ".png"))


def move_img():
    if numEty.get() == '':
        messagebox("erro", "Indica o Id")
    else:
        if not os.path.exists('./IMG/' + numEty.get() + ".png"):
            messagebox("erro", "Cria a imagem primeiro")
        else:
            global photo_04, img_04
            filepath = filedialog.askopenfilename()
            if filepath == '':
                messagebox("erro", "Não seleccionou nenhum documento")
            else:
                img_04 = Image.open(filepath)
                photo_04 = ImageTk.PhotoImage(img_04)
                lbl_Fm_Img['image'] = photo_04
                img_04.save(os.path.join("./IMG/", numEty.get() + ".png"))


def dele_img():
    if numEty.get() == '':
        messagebox("erro", "Indica o Id")
    else:
        if not os.path.exists('./IMG/' + numEty.get() + ".png"):
            messagebox("erro", "Não existe a imagem")
        else:
            os.remove('./IMG/' + numEty.get() + ".png")
            messagebox("delete", "Imagem eliminado")



# opecoe da janela
root = Tk()  # Criar um objecto de janela principal, chamado root
root.title('Menu')  # Titlo de janela
root.geometry('1000x800')  # Tamanho de janela
root.iconbitmap('./SB-Vidros.ico')  # icon de janela
root["background"] = "#5ab100"  # Cor de fundo de janela

img_01 = Image.open('SB-V.png')
photo_01 = ImageTk.PhotoImage(img_01)

# Componentes da janela
title = Label(root, text='Informação pessoal', font=('Calibri', 50, 'bold italic'), bg="white",
              fg="red", relief="raised", padx=5)
numLbl = Label(root, text='Insere o Id que procure:')
procBtn = Button(root, text='Procurar', command=procurar)
numEty = Entry(root)
btn_Clear = Button(root, text='Limpar', command=clear)
btn_Save = Button(root, text='Modificar', command=salvar)
btn_Create = Button(root, text='Criar uma nova', command=create)
btn_Del = Button(root, text='Apagar', command=delete_data)
btn_QRC = Button(root, text='Crear QRCode', command=create_qrc)
lbl_img = Label(root, image=photo_01)

fm = Frame(root)
lbl_Id = Label(fm, text='Id:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
lbl_Id_V = Label(fm, relief="sunken", borderwidth=5, anchor='w')
lbl_Name = Label(fm, text='Nome:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_Name = Text(fm)
lbl_tele = Label(fm, text='Telemovel:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_tele = Text(fm)
lbl_telf = Label(fm, text='Telefone:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_telf = Text(fm)
lbl_email = Label(fm, text='Email:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_email = Text(fm)
lbl_morada = Label(fm, text='Morada:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_morada = Text(fm)
lbl_empr = Label(fm, text='Empresa:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_empr = Text(fm)
lbl_ocup = Label(fm, text='Ocupação:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_ocup = Text(fm)
lbl_local = Label(fm, text='Local:', font=('Calibri', 20, 'bold italic'), anchor='e', justify='left')
txt_local = Text(fm)

btn_Crate_Img = Button(fm, text='Carregar imagem', command=create_img)
btn_Move_Img = Button(fm, text='Modificar imagem', command=move_img)
btn_Del_Img = Button(fm, text='Eleminar imagem', command=dele_img)
lbl_Fm_Img = Label(fm, image=photo_01)

# Componentes por na janela
title.place(x=40, y=40, width=600, height=80)

numLbl.place(x=40, y=125, width=170, height=20)
numEty.place(x=215, y=125, width=175, height=20)

procBtn.place(x=440, y=125, width=200, height=20)
btn_Clear.place(x=440, y=150, width=200, height=20)
btn_Save.place(x=440, y=175, width=200, height=20)
btn_Create.place(x=440, y=200, width=200, height=20)
btn_Del.place(x=440, y=225, width=200, height=20)
btn_QRC.place(x=440, y=250, width=200, height=20)

lbl_img.place(x=690, y=40)

fm.place(x=40, y=150, width=375, height=600)

lbl_Id.place(x=10, y=5, width=130, height=20)
lbl_Id_V.place(x=155, y=5, width=200, height=25)
lbl_Name.place(x=10, y=40, width=130, height=20)
txt_Name.place(x=155, y=40, width=200, height=20)
lbl_tele.place(x=10, y=80, width=130, height=20)
txt_tele.place(x=155, y=80, width=200, height=20)
lbl_telf.place(x=10, y=120, width=130, height=20)
txt_telf.place(x=155, y=120, width=200, height=20)
lbl_email.place(x=10, y=160, width=130, height=20)
txt_email.place(x=155, y=160, width=200, height=20)
lbl_empr.place(x=10, y=200, width=130, height=20)
txt_empr.place(x=155, y=200, width=200, height=20)
lbl_ocup.place(x=10, y=240, width=130, height=20)
txt_ocup.place(x=155, y=240, width=200, height=20)
lbl_local.place(x=10, y=280, width=130, height=20)
txt_local.place(x=155, y=280, width=200, height=20)
lbl_Fm_Img.place(x=80, y=320, width=220, height=220)
btn_Crate_Img.place(x=10, y=560, width=115, height=20)
btn_Move_Img.place(x=130, y=560, width=115, height=20)
btn_Del_Img.place(x=250, y=560, width=115, height=20)

# Janela em exposição
root.mainloop()
# guardar os dados no excel
wb.save('dados.xlsx')
